from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
import os.path
import time


class Excel(object):
    def __init__(self,
                 nsheet=0,
                 tm_url_tag='天猫网址',
                 tm_price_tag='天猫价',
                 jd_url_tag='京东网址',
                 jd_price_tag='京东价',
                 sku_tag='SKU',
                 name_tag='商品名称'):
        self.path = self.__get_path()
        self.__open(nsheet)
        self.fields = self.__get_fields()
        self.index = self.__get_index(tm_url_tag, tm_price_tag, jd_url_tag, jd_price_tag, sku_tag,
                                    name_tag)
        # 存储异常商品所在行
        self.err_rows = set()
        # 设置字体
        self.__ft = Font(color=colors.RED)

    def __get_path(self):
        '''获取excel文件目录,路径不要带空格'''
        p = input('把表格拖进这个窗口，后按下回车\n')
        # 去除路径末尾的空格
        while p.endswith(' '):
            p = p[:-1]
        # 保存路径
        path = {'source': p, 'dirname': os.path.dirname(p), 'basename': os.path.basename(p)}
        return path

    def __open(self, n):
        # 加载excel文件中的第n个sheet
        self.wb = load_workbook(self.path['source'])
        self.ws = self.wb.worksheets[n]
        # 统计总共多少行多少列
        self.nrows = self.ws.max_row
        self.ncols = self.ws.max_column
        # 商品计数
        self.items_counting = 1
        self.items_total = self.nrows - 1
        print('加载成功，共有{}个商品'.format(self.items_total))

    def __get_fields(self):
        '''获取第一行表头的字段'''
        fields = []
        for i in range(self.ncols):
            fields.append(self.ws.cell(row=1, column=i + 1).value)
        return fields

    def __get_index(self, tm_url_tag, tm_price_tag, jd_url_tag, jd_price_tag, sku_tag, name_tag):
        # 确定所需字段在第几列
        index = {}
        index['tm_url'] = self.fields.index(tm_url_tag)
        index['tm_price'] = self.fields.index(tm_price_tag)
        index['jd_url'] = self.fields.index(jd_url_tag)
        index['jd_price'] = self.fields.index(jd_price_tag)
        index['sku'] = self.fields.index(sku_tag)
        index['name'] = self.fields.index(name_tag)
        return index

    def save(self):
        now = time.strftime('%m%d')
        self.path['filename'] = now + '已抓取_' + self.path['basename']
        self.path['target'] = os.path.join(self.path['dirname'], self.path['filename'])
        # 保存新表格
        self.wb.save(self.path['target'])
        print('''
        *************************************************************************
        完成!已保存到 {}
        *************************************************************************
        '''.format(self.path['target']))

    def fetch_tm_url_by_row(self, row):
        '''获取指定行的天猫链接'''
        cell = self.ws.cell(row=row + 1, column=self.index['tm_url'] + 1)
        value = cell.value
        return value

    def set_tm_price_by_row(self, row, value):
        '''写入指定行的天猫价格'''
        cell = self.ws.cell(row=row + 1, column=self.index['tm_price'] + 1)
        cell.value = value
        # 如果值是一个字符串(非价格数字)，把字体改成红色，并且记录错误行数
        if isinstance(value, str):
            cell.font = self.__ft
            self.err_rows.add(row)

    def fetch_jd_url_by_row(self, row):
        '''获取指定行的京东链接'''
        cell = self.ws.cell(row=row + 1, column=self.index['jd_url'] + 1)
        value = cell.value
        return value

    def set_jd_price_by_row(self, row, value):
        '''写入指定行的京东价格'''
        cell = self.ws.cell(row=row + 1, column=self.index['jd_price'] + 1)
        cell.value = value
        # 如果值是一个字符串(非价格数字)，把字体改成红色，并且记录错误行数
        if isinstance(value, str):
            cell.font = self.__ft
            self.err_rows.add(row)

    def fetch_sku_by_row(self, row):
        '''获取指定行的sku'''
        cell = self.ws.cell(row=row + 1, column=self.index['sku'] + 1)
        value = cell.value
        return value

    def fetch_name_by_row(self, row):
        '''获取指定行的商品名称'''
        cell = self.ws.cell(row=row + 1, column=self.index['name'] + 1)
        value = cell.value
        return value


# 测试
if __name__ == '__main__':
    test = Excel()
    print(test.path)
    print(test.fields)
    print(test.index)
    print(test.fetch_tm_url(3))
    print(test.fetch_tm_url(7))