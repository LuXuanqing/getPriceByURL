from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from tmall import getTmallPrice
from jd import getJdPrice
import os.path
import time

# 获取文件路径
path = {}
path['source'] = input('把表格拖进这个窗口，然后按下回车\n')
# 去除路径末尾的空格
while path['source'].endswith(' '):
    path['source'] = path['source'][:-1]

# 加载excel文件中的第一个sheet
wb = load_workbook(path['source'])
ws = wb.worksheets[0]
# 统计总共多少行多少列
nrows = ws.max_row
ncols = ws.max_column
# print('there are {} rows and {} columns'.format(nrows, ncols))

# 获取第一行表头的所有字段
fields = []
for i in range(ncols):
    fields.append(ws.cell(row=1,column=i+1).value)
# print(fields)
# 确定所需字段在第几列
t_url_tag = '天猫网址'
t_price_tag = '天猫价'
jd_url_tag = '京东网址'
jd_price_tag = '京东价'
sku_tag = 'SKU'
name_tag = '商品名称'
t_url_index = fields.index(t_url_tag)
t_price_index = fields.index(t_price_tag)
jd_url_index = fields.index(jd_url_tag)
jd_price_index = fields.index(jd_price_tag)
sku_index = fields.index(sku_tag)
name_index = fields.index(name_tag)

# 商品计数
items_counting = 1
items_total = nrows - 1
print('共有{}个商品'.format(items_total))

# 存储异常商品所在行
errs = set()

# 设置异常数据所在单元格样式
ft = Font(color=colors.RED)

# 遍历所有行中的这四个字段
for row in range(1, nrows):
    # 获得该行所需的单元格
    t_url_cell = ws.cell(row=row+1, column=t_url_index+1)
    t_price_cell = ws.cell(row=row+1, column=t_price_index+1)
    jd_url_cell = ws.cell(row=row+1, column=jd_url_index+1)
    jd_price_cell = ws.cell(row=row+1, column=jd_price_index+1)
    sku_cell = ws.cell(row=row+1, column=sku_index+1)
    name_cell = ws.cell(row=row+1, column=name_index+1)
    # 打印查询进度信息
    print('({}/{})正在查询:({}){}'.format(items_counting, items_total, sku_cell.value, name_cell.value))
    items_counting += 1
    # 查询并修改价格
    if t_url_cell.value:
        t_price_cell.value = getTmallPrice(t_url_cell.value)
        if t_price_cell.value < 0:
            t_price_cell.font = ft
            errs.add(row)
    if jd_url_cell.value:
        jd_price_cell.value = getJdPrice(jd_url_cell.value)
        if jd_price_cell.value < 0:
            jd_price_cell.font = ft
            errs.add(row)

# 设置保存路径以及文件名
path['dirname'] = os.path.dirname(path['source'])
path['basename'] = os.path.basename(path['source'])
now = time.strftime('%m%d')
path['filename'] = now + '已抓取_' + path['basename']
path['target'] = os.path.join(path['dirname'], path['filename'])
# print(path)

# 保存新表格
wb.save(path['target'])
print('''
*************************************************************************
完成!已保存到原文件下，{}
*************************************************************************
'''.format(path['filename']))

# 打印异常商品信息
if len(errs) > 0:
    key = input('{}个商品的价格查询有异常，按回车查看，输入q退出\n'.format(len(errs)))
    if key == 'q':
        quit()
    elif key == '':
        err_rows = list(errs)

        def printItem(row):
            '''输出指定行的商品sku和名称。
            '''
            sku = ws.cell(row=row+1, column=sku_index+1).value
            name = ws.cell(row=row+1, column=name_index+1).value
            print('({}){}'.format(sku, name))

        print('\n+++++++++++++++以下商品价格查询有异常，记得检查一下哦+++++++++++++++')
        for row in err_rows:
            printItem(row)
        print('+++++++++++++++以上商品价格查询有异常，记得检查一下哦+++++++++++++++\n')

    key = input('错误代码的相关信息，按回车查看，输入q退出\n')
    if key == 'q':
        quit()
    elif key == '':
        print('''
        -101：商品的url错误
        -102：无法获取商品id
        -103：发送的HTTP请求收不到返回数据
        -104：服务器没有返回正确数据
        ''')
# 退出程序
input('按回车退出程序')